import logging
import json
import math
import os
import inspect
import copy
from ckan.plugins import toolkit
from ckan.lib.helpers import url_for, get_site_protocol_and_host
from ckantoolkit import config
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from flask import Response
from datetime import datetime
from ckan.common import _

DATASETS_PER_PAGE = 100
wrong_page_exception = toolkit.ValidationError(
    'Page param must be a positive integer starting in 1')
from dateutil.parser import parse as dateutil_parse

log = logging.getLogger(__name__)

@toolkit.side_effect_free
def hutemplate_export_excel(context, data_dict):
    toolkit.check_access('hutemplate_export_excel', context)
    
    config = load_config()
    
    book = Workbook()
    ws = book.active
    ws.title = _('Datasets, resources') # 'Adatkészlet, adatelérés'
    # Worksheet header
    row = 1
    col_max_char_lengths = {}
    _create_dataset_header(ws, config, row, col_max_char_lengths)
    
    # Rows
    row += 2
    
    cache_map = {}
    data_dict = {
        'page': 1
    }
    query = _search_ckan_datasets(context, data_dict, 1)
    total_count = query['count']

    n = int(config.get('ckanext.dcat.datasets_per_page', DATASETS_PER_PAGE))
    total_pages =  math.ceil(total_count / n)

    for page in range(1,total_pages+1):
        data_dict = {
            'page': page
        }
        query = _search_ckan_datasets(context, data_dict, n)
        dataset_dicts = query['results']
        for data_row in dataset_dicts:
            row =_create_dataset_row(context, ws, config, row, data_row, col_max_char_lengths, cache_map)

    for i, j in enumerate(col_max_char_lengths,1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = col_max_char_lengths[i]

    ws2 = book.create_sheet(_('Organization')) # 'Szervezet'
    row = 1
    col_max_char_lengths = {1: 0, 2: 0}

    _create_organization_header(ws2, row, col_max_char_lengths)
    org_list = toolkit.get_action('organization_list')(context, {'all_fields':True})
    row += 1
    for org in org_list:
        row = _create_organization_row(ws2, row, org, col_max_char_lengths)

    for i, j in enumerate(col_max_char_lengths,1):  # ,1 to start at 1
        ws2.column_dimensions[get_column_letter(i)].width = col_max_char_lengths[i]

    return Response(
        save_virtual_workbook(book),
        headers={
            'Content-Disposition': 'attachment; filename='+_('Datasets')+'.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )

def _create_dataset_row(context, ws, config, row, data_row, col_max_char_lengths, cache_map):
    mapping = config['mapping']
    col_groups = mapping.keys()
    flattened_data_rows = _flattend_data_row(data_row)
    for flattened_data in flattened_data_rows:
        col = 1
        for col_group in col_groups:
            values_in_order = mapping[col_group].values()
            for field_value in values_in_order:
                field_arr = field_value.split('.')
                obj = field_arr[0]
                field = field_arr[1]
                if obj == 'package':
                    _add_package_field(context, ws, row, col, field, flattened_data, col_max_char_lengths, cache_map)
                elif obj == 'resource':
                    _add_resource_field(context, ws, row, col, field, flattened_data, col_max_char_lengths, cache_map)
                col += 1
        row += 1
    return row

def _create_organization_row(ws, row, org, col_max_char_lengths):
    col = 1
    cell = ws.cell(row=row, column=col)
    cell.value = org['id']
    col_max_char_lengths[col] = max(col_max_char_lengths[col], len(org['id']))
    
    col = 2
    cell = ws.cell(row=row, column=col)
    cell.value = org['display_name']
    col_max_char_lengths[col] = max(col_max_char_lengths[col], len(org['display_name']))
    row += 1

    return row

def _add_package_field(context, ws, row, col, field, data_row, col_max_char_lengths, cache_map):
    value = None
    if field in data_row:
        value = data_row[field]
        
    if field == 'groups' or field == 'tags':
        if value is not None:
            value = ', '.join([v['display_name'] for v in value])
    #multiple text format
    elif field == 'documentation' or field == 'conforms_to' or field == 'alternate_identifier' or field == 'source':
        if value is not None:
            value = ', '.join(value)
    elif field == 'organization_title':
        value = data_row['organization']['title']
    elif field == 'creator_user_name':
        value = _get_username_for_id(context, cache_map, data_row['creator_user_id'])
    elif field == 'dcat_type' or field == 'access_rights' or field == 'temporal_resolution' or field == 'frequency':
        value = _get_field_translation(context, cache_map, field, value)
    elif field == 'name':
        value = get_site_protocol_and_host()[0]+'://'+get_site_protocol_and_host()[1]+url_for('dataset.read', id=value)
    
    if value is None:
        if col not in col_max_char_lengths:
            col_max_char_lengths[col] = 0
        return

    cell = ws.cell(row=row, column=col)
    cell.value = value
    if col not in col_max_char_lengths:
        col_max_char_lengths[col] = 0

    col_max_char_lengths[col] = max(col_max_char_lengths[col], len(value))

def _add_resource_field(context, ws, row, col, field, data_row, col_max_char_lengths, cache_map):
    value = None
    if field in data_row['resource']:
        value = data_row['resource'][field]
    
    if field == 'conforms_to':
        if value is not None:
            value = ', '.join(value)
    elif field == 'status' or field == 'temporal_resolution':
        value = _get_field_translation(context, cache_map, field, value)
    elif field == 'url_type':
        value = "Igen" if value=='upload' else "Nem"
    elif field == 'metadata_modified':
        if value is not None:
            value = value[0:10]

    if value is None:
        if col not in col_max_char_lengths:
            col_max_char_lengths[col] = 0
        return

    cell = ws.cell(row=row, column=col)
    cell.value = value
    if col not in col_max_char_lengths:
        col_max_char_lengths[col] = 0

    col_max_char_lengths[col] = max(col_max_char_lengths[col], len(value))

def _get_field_translation(context, cache_map, field, value):
    if value is None:
        return None
    if 'translation' not in cache_map:
        cache_map['translation'] = {}
    if value in cache_map['translation']:
        return cache_map['translation'][value]
    translations = toolkit.get_action('term_translation_show')(context,{'terms': value, 'lang_codes': ("hu")})
    if len(translations) > 0:
        cache_map['translation'][value] = translations[0]['term_translation']
    else:
        cache_map['translation'][value] = value
    return cache_map['translation'][value]

def _get_username_for_id(context, cache_map, user_id):
    if 'user_name' not in cache_map:
        cache_map['user_name'] = {}
    if user_id in cache_map['user_name']:
        return cache_map['user_name'][user_id]
    user_data = toolkit.get_action('user_show')(context, {'id': user_id})
    cache_map['user_name'][user_id] = user_data['fullname']
    return user_data['fullname']

def _flattend_data_row(data_row):
    returned = []
    for resource in data_row['resources']:
        row = copy.deepcopy(data_row)
        row['resources'] = None
        row['resource'] = resource
        returned.append(row)
    return returned

def _create_organization_header(ws, row, col_max_char_lengths):
    col = 1
    cell = ws.cell(row=row, column=col)
    cell.value = 'Szervezet ID'
    cell.font = Font(bold=True)
    col_max_char_lengths[col] = max(col_max_char_lengths[col], len('Szervezet ID'))

    col += 1
    cell = ws.cell(row=row, column=col)
    cell.value = 'Szervezet neve'
    cell.font = Font(bold=True)
    col_max_char_lengths[col] = max(col_max_char_lengths[col], len('Szervezet neve'))

def _create_dataset_header(ws, config, row, col_max_char_lengths):
    mapping = config['mapping']
    col_groups = mapping.keys() #Adatkészlet, Adatelérés
    
    col = 1
    for col_group in col_groups:
        col_names_in_order = mapping[col_group]
        group_cell_len = len(col_names_in_order)
        cell = ws.cell(row=row, column=col)
        cell.value = col_group
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+group_cell_len-1)
        col += group_cell_len
    
    row += 1
    col = 1
    for col_group in col_groups:
        col_names_in_order = mapping[col_group].keys()
        for col_name in col_names_in_order:
            cell = ws.cell(row=row, column=col)
            cell.value = col_name
            cell.font = Font(bold=True)
            if col not in col_max_char_lengths:
                col_max_char_lengths[col] = 0

            col_max_char_lengths[col] = max(col_max_char_lengths[col], len(col_name))
            col += 1

def _search_ckan_datasets(context, data_dict, n):

    
    page = data_dict.get('page', 1) or 1

    try:
        page = int(page)
        if page < 1:
            raise wrong_page_exception
    except ValueError:
        raise wrong_page_exception

    search_data_dict = {
        'rows': n,
        'start': n * (page - 1),
        'sort': 'metadata_modified desc',
    }

    search_data_dict['q'] = data_dict.get('q', '*:*')
    search_data_dict['fq'] = data_dict.get('fq')
    search_data_dict['fq_list'] = []

    # Exclude certain dataset types
    search_data_dict['fq_list'].append('-dataset_type:harvest')
    search_data_dict['fq_list'].append('-dataset_type:showcase')

    query = toolkit.get_action('package_search')(context, search_data_dict)

    return query

@toolkit.auth_allow_anonymous_access
def hutemplate_auth(context, data_dict):
    '''
    All users can access endpoints by default
    '''
    return {'success': True}

def load_config(): 
    # Opening JSON file
    m = __import__('ckanext.hutemplate', fromlist=[''])
    p = os.path.join(os.path.dirname(inspect.getfile(m)), 'excel_config.json')

    f = open(p, encoding="utf-8")
    config = json.load(f)
    # Closing file
    f.close()
    return config